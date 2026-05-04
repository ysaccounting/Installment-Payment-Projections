"""
Y&S Tickets — Installment Projections  (Streamlit app)
Drop in one or more General Ledger files, download formatted reports.
"""

import io, re
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Installment Projections", page_icon="🎟️", layout="centered")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:wght@300;400;500;600&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;}
.stApp{background:#F5F3EF;}
#MainMenu,footer,header{visibility:hidden;}
.hero{text-align:center;padding:3rem 0 2rem;}
.hero h1{font-family:'DM Serif Display',serif;font-size:2.8rem;color:#1F3864;margin:0;line-height:1.1;letter-spacing:-0.5px;}
.hero h1 em{font-style:italic;color:#2E5EAA;}
.hero p{font-size:1.05rem;color:#666;margin-top:.6rem;font-weight:300;}
.upload-card{background:white;border-radius:16px;padding:2.5rem 2rem;box-shadow:0 2px 20px rgba(31,56,100,.08);border:1.5px solid #E8E4DC;margin:1.5rem 0;}
.step-label{font-size:.7rem;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:#2E5EAA;margin-bottom:.5rem;}
.success-box{background:#EEF4FB;border-left:4px solid #2E5EAA;border-radius:8px;padding:1rem 1.2rem;margin:1rem 0;}
.success-box .file-title{font-family:'DM Serif Display',serif;font-size:1.1rem;color:#1F3864;margin:0;}
.success-box .file-meta{font-size:.85rem;color:#555;margin:.2rem 0 0;}
.stats-row{display:flex;gap:1rem;margin:1rem 0;}
.stat-card{flex:1;background:#F5F3EF;border-radius:10px;padding:.9rem 1rem;text-align:center;border:1px solid #E8E4DC;}
.stat-card .stat-num{font-family:'DM Serif Display',serif;font-size:1.6rem;color:#1F3864;line-height:1;}
.stat-card .stat-lbl{font-size:.72rem;color:#888;text-transform:uppercase;letter-spacing:1px;margin-top:.2rem;}
.divider{border:none;border-top:1px solid #E8E4DC;margin:1.5rem 0;}
[data-testid="stFileUploader"]{background:#FAFAF8;border:2px dashed #C5D8F0;border-radius:12px;padding:1rem;}
[data-testid="stFileUploader"]:hover{border-color:#2E5EAA;}
[data-testid="stDownloadButton"]>button{background:#1F3864!important;color:white!important;border:none!important;border-radius:10px!important;padding:.7rem 2rem!important;font-family:'DM Sans',sans-serif!important;font-weight:600!important;font-size:.95rem!important;width:100%!important;transition:background .2s!important;}
[data-testid="stDownloadButton"]>button:hover{background:#2E5EAA!important;}
.err-box{background:#FFF0F0;border-left:4px solid #C00000;border-radius:8px;padding:1rem 1.2rem;color:#C00000;font-size:.9rem;}
.footer{text-align:center;color:#aaa;font-size:.78rem;padding:2rem 0 1rem;}
</style>
""", unsafe_allow_html=True)

# ── CONSTANTS ─────────────────────────────────────────────────────────────────

PRIMARY_ACCOUNTS = {'Slash','Divvy Credit','Divvy Prefund','Wex Credit','Wex Prefund','Global Rewards'}
EXCLUDE_ACCOUNTS = {'Clearing Account','Accounts Payable'}
DARK_BLUE='1F3864'; MED_BLUE='2E5EAA'; TEAL='1F6B75'; GREEN='1F6B3B'
ALT_ROW='EEF4FB'; ALT_OTHER='F0F7F4'; ALT_TEAM='F0F7F0'; WHITE='FFFFFF'

# ── VALIDATION ────────────────────────────────────────────────────────────────

def validate_gl(df):
    ncols = df.shape[1]
    if ncols < 9:
        return False, (f"This looks like a previously generated report ({ncols} columns). "
                       "Please upload the raw QuickBooks GL export.")
    for col in [8,9]:
        if col < ncols and pd.to_numeric(df[col],errors='coerce').notna().sum() > 5:
            return True, None
    return False, "This doesn't look like a General Ledger export."

# ── FORMAT DETECTION ─────────────────────────────────────────────────────────

def detect_format(df):
    if 9 not in df.columns or 8 not in df.columns: return 'old'
    mask = df[9].notna() & df[0].isna() & df[1].notna() & \
           ~df[1].astype(str).str.contains('Beginning Balance|Transaction date',case=False,na=False)
    sample = df[mask][8].dropna()
    if len(sample)>0 and sample.nunique()<=20 and (len(sample)/max(sample.nunique(),1))>5:
        try: pd.to_numeric(sample,errors='raise')
        except (ValueError,TypeError): return 'consolidated'
    return 'new' if str(df.iloc[1,0]).strip()=='Transaction Report' else 'old'

# ── DATE HELPERS ──────────────────────────────────────────────────────────────

def ordinal(n):
    return {1:'st',2:'nd',3:'rd'}.get(n%10,'th') if n not in (11,12,13) else 'th'

def next_month_date_label(date_range_str):
    actual = date_range_str.strip()
    m = re.search(r'(\w+)\s+(\d+)\w*[-–](\d+)\w*,?\s*(\d{4})',actual)
    if m:
        month_str,start_day,end_day,year_str = m.group(1),int(m.group(2)),int(m.group(3)),m.group(4)
        nxt = pd.to_datetime(f'{month_str} 1 {year_str}')+pd.offsets.MonthBegin(1)
        return f'{nxt.strftime("%B")} {start_day}{ordinal(start_day)}-{end_day}{ordinal(end_day)}', actual
    mm=re.search(r'(\w+)\s+(\d+)',actual); ym=re.search(r'(\d{4})',actual)
    month_str=mm.group(1)
    start_day=int(mm.group(2)) if ym and mm.group(2)!=ym.group(1) else 1
    year_str=ym.group(1) if ym else None
    nxt=pd.to_datetime(f'{month_str} 1 {year_str}')+pd.offsets.MonthBegin(1)
    return f'{nxt.strftime("%B")} {start_day}{ordinal(start_day)}', actual

# ── METADATA ─────────────────────────────────────────────────────────────────

def smart_title(s):
    ALLCAPS={'llc','tl','ys','yss','gl','mls','nba','nhl','mlb','nfl','tc','dep','lp','inc','ltd','kg','sp','yskg','ysp','ysm','gr'}
    result=[]
    for w in s.split():
        if '&' in w: result.append(w.upper() if len(w)<=3 else w.title())
        elif w.lower() in ALLCAPS: result.append(w.upper())
        else: result.append(w.capitalize())
    return ' '.join(result)

def build_meta(df,fmt):
    if fmt=='consolidated':
        name='All Companies'; date_range_str=str(df.iloc[1].dropna().tolist()[0]).strip()
    elif fmt=='old':
        name=str(df.iloc[1].dropna().tolist()[0]).strip(); date_range_str=str(df.iloc[2].dropna().tolist()[0]).strip()
    else:
        name=str(df.iloc[0].dropna().tolist()[0]).strip(); date_range_str=str(df.iloc[2].dropna().tolist()[0]).strip()
    next_label,actual_label=next_month_date_label(date_range_str)
    company_title=smart_title(name)
    report_title=f'{company_title} - Installment Projections - {next_label}'
    return company_title,actual_label,next_label,report_title,f'{report_title}.xlsx'

# ── ACCOUNT RELABELING ────────────────────────────────────────────────────────

def relabel(account):
    a=str(account).strip(); al=a.lower()
    if 'slash plat' in al or al.startswith('sp') or ' sp' in al: return 'Slash'
    if 'divvy cr' in al or al in ('divvy (credit)','divvy credit'): return 'Divvy Credit'
    if 'divvy pf' in al or al=='divvy (prefund)': return 'Divvy Prefund'
    if 'wex (prefund)' in al or 'wex prefund' in al: return 'Wex Prefund'
    if 'wex cr' in al or 'wex (credit' in al: return 'Wex Credit'
    if 'global reward' in al or al.startswith('gr ') or al=='gr': return 'Global Rewards'
    return a

# ── DATA LOADING ──────────────────────────────────────────────────────────────

def load_transactions(df,fmt):
    if fmt=='old':
        amt_col=8;acct_col=7;type_col=2;name_col=4;desc_col=5;date_col=1;comp_col=None
        mask=df[amt_col].notna()&(df[amt_col]!='Amount')&df[0].isna()
    elif fmt=='new':
        amt_col=9;acct_col=8;type_col=2;name_col=5;desc_col=6;date_col=1;comp_col=None
        mask=(df[amt_col].notna()&df[0].isna()&df[date_col].notna()&
              ~df[date_col].astype(str).str.contains('Beginning Balance|Date',case=False,na=False))
    else:
        amt_col=9;acct_col=7;type_col=2;name_col=4;desc_col=5;date_col=1;comp_col=8
        mask=(df[amt_col].notna()&df[0].isna()&df[date_col].notna()&
              ~df[date_col].astype(str).str.contains('Beginning Balance|Transaction date',case=False,na=False))
    tx=df[mask].copy(); tx[amt_col]=pd.to_numeric(tx[amt_col],errors='coerce'); tx=tx.dropna(subset=[amt_col])
    tx['labeled']=tx[acct_col].apply(relabel)
    tx=tx[tx[type_col].notna()&(tx[type_col].astype(str).str.strip()!='')]
    tx=tx[~tx['labeled'].isin(EXCLUDE_ACCOUNTS)]; tx=tx[~tx[acct_col].isin(EXCLUDE_ACCOUNTS)]
    cols=[date_col,type_col,name_col,desc_col,'labeled',amt_col]; names=['Date','Type','Name','Description','Account','Amount']
    if comp_col is not None: cols.append(comp_col); names.append('Company')
    result=tx[cols].copy(); result.columns=names; result['Date']=pd.to_datetime(result['Date'],errors='coerce')
    return result[result['Amount']!=0]

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────

def hfont(sz=11): return Font(name='Arial',bold=True,color=WHITE,size=sz)
def cfont(sz=10): return Font(name='Arial',size=sz)
def tborder():
    s=Side(style='thin',color='BFBFBF'); return Border(left=s,right=s,top=s,bottom=s)
def bold_left_border():
    m=Side(style='medium',color='BFBFBF'); t=Side(style='thin',color='BFBFBF')
    return Border(left=m,right=t,top=t,bottom=t)

def write_title(ws,report_title,actual_date_range,ncols):
    sp=get_column_letter(ncols)
    ws.merge_cells(f'A1:{sp}1'); c=ws['A1']; c.value=report_title
    c.font=Font(name='Arial',bold=True,color=WHITE,size=13); c.fill=PatternFill('solid',fgColor=DARK_BLUE)
    c.alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[1].height=30
    ws.merge_cells(f'A2:{sp}2'); c=ws['A2']; c.value=actual_date_range
    c.font=Font(name='Arial',italic=True,color=WHITE,size=10); c.fill=PatternFill('solid',fgColor=MED_BLUE)
    c.alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[2].height=18; ws.row_dimensions[3].height=8

def write_sec_hdr(ws,row,label,color,c1,c2):
    ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
    c=ws.cell(row=row,column=c1,value=label)
    c.font=Font(name='Arial',bold=True,color=WHITE,size=10); c.fill=PatternFill('solid',fgColor=color)
    c.alignment=Alignment(horizontal='left',vertical='center',indent=1); c.border=tborder(); ws.row_dimensions[row].height=20

def write_col_hdrs(ws,row,hdrs,bg,c1=1):
    for i,h in enumerate(hdrs):
        c=ws.cell(row=row,column=c1+i,value=h)
        c.font=Font(name='Arial',bold=True,color=WHITE,size=10); c.fill=PatternFill('solid',fgColor=bg)
        c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tborder()
    ws.row_dimensions[row].height=20

# ── ACCOUNT SUMMARY (cols 1-4) ────────────────────────────────────────────────

def write_account_summary(ws,tx,start_row):
    acct=tx.groupby('Account')['Amount'].agg(['sum','count']).reset_index()
    acct.columns=['Account','Total Spent','Transactions']
    acct=acct.sort_values('Total Spent',ascending=False).reset_index(drop=True)
    prim=acct[acct['Account'].isin(PRIMARY_ACCOUNTS)].sort_values('Total Spent',ascending=False).reset_index(drop=True)
    oth=acct[~acct['Account'].isin(PRIMARY_ACCOUNTS)].sort_values('Total Spent',ascending=False).reset_index(drop=True)
    grand_total=float(tx['Amount'].sum())

    r=start_row
    write_sec_hdr(ws,r,'▸  Primary Accounts',MED_BLUE,1,4); r+=1
    write_col_hdrs(ws,r,['Account','Total Spent ($)','# Trans','% Of Total'],'4472C4',1); r+=1

    for i,row in prim.iterrows():
        fill=ALT_ROW if i%2==0 else WHITE; pct=row['Total Spent']/grand_total if grand_total else 0
        for col,(val,fmt,aln) in enumerate(zip([row['Account'],row['Total Spent'],int(row['Transactions']),pct],
                                               [None,'$#,##0.00','#,##0','0.0%'],['left','center','center','center']),1):
            c=ws.cell(row=r,column=col,value=val); c.font=cfont(); c.fill=PatternFill('solid',fgColor=fill)
            c.border=tborder(); c.alignment=Alignment(horizontal=aln,vertical='center')
            if fmt: c.number_format=fmt
        ws.row_dimensions[r].height=18; r+=1

    p_total=float(prim['Total Spent'].sum()); p_pct=p_total/grand_total if grand_total else 0
    for col,(val,fmt,aln) in enumerate(zip(['Subtotal — Primary',p_total,int(prim['Transactions'].sum()),p_pct],
                                           [None,'$#,##0.00','#,##0','0.0%'],['left','center','center','center']),1):
        c=ws.cell(row=r,column=col,value=val); c.font=Font(name='Arial',bold=True,color=WHITE,size=10)
        c.fill=PatternFill('solid',fgColor=MED_BLUE); c.border=tborder()
        c.alignment=Alignment(horizontal=aln,vertical='center')
        if fmt: c.number_format=fmt
    ws.row_dimensions[r].height=20; r+=1; ws.row_dimensions[r].height=8; r+=1

    write_sec_hdr(ws,r,'▸  Other Accounts',TEAL,1,4); r+=1
    write_col_hdrs(ws,r,['Account','Total Spent ($)','# Trans','% Of Total'],'2E8B8F',1); r+=1

    for i,row in oth.iterrows():
        fill=ALT_OTHER if i%2==0 else WHITE; pct=row['Total Spent']/grand_total if grand_total else 0
        for col,(val,fmt,aln) in enumerate(zip([row['Account'],row['Total Spent'],int(row['Transactions']),pct],
                                               [None,'$#,##0.00','#,##0','0.0%'],['left','center','center','center']),1):
            c=ws.cell(row=r,column=col,value=val); c.font=cfont(); c.fill=PatternFill('solid',fgColor=fill)
            c.border=tborder(); c.alignment=Alignment(horizontal=aln,vertical='center')
            if fmt: c.number_format=fmt
        ws.row_dimensions[r].height=18; r+=1

    o_total=float(oth['Total Spent'].sum()); o_pct=o_total/grand_total if grand_total else 0
    for col,(val,fmt,aln) in enumerate(zip(['Subtotal — Other',o_total,int(oth['Transactions'].sum()),o_pct],
                                           [None,'$#,##0.00','#,##0','0.0%'],['left','center','center','center']),1):
        c=ws.cell(row=r,column=col,value=val); c.font=Font(name='Arial',bold=True,color=WHITE,size=10)
        c.fill=PatternFill('solid',fgColor=TEAL); c.border=tborder()
        c.alignment=Alignment(horizontal=aln,vertical='center')
        if fmt: c.number_format=fmt
    ws.row_dimensions[r].height=20; r+=1; ws.row_dimensions[r].height=8; r+=1

    for col,(val,fmt,aln) in enumerate(zip(['Grand Total',grand_total,int(acct['Transactions'].sum()),1.0],
                                           [None,'$#,##0.00','#,##0','0.0%'],['left','center','center','center']),1):
        c=ws.cell(row=r,column=col,value=val); c.font=Font(name='Arial',bold=True,color=WHITE,size=11)
        c.fill=PatternFill('solid',fgColor=DARK_BLUE); c.border=tborder()
        c.alignment=Alignment(horizontal=aln,vertical='center')
        if fmt: c.number_format=fmt
    ws.row_dimensions[r].height=24; r+=1
    return r

# ── TEAM SUMMARY (cols 6-9) ───────────────────────────────────────────────────

def write_team_summary(ws,tx,start_row):
    team=tx.groupby('Name')['Amount'].agg(['sum','count']).reset_index()
    team.columns=['Team','Total Spent','Transactions']
    team=team.sort_values('Total Spent',ascending=False).reset_index(drop=True)
    grand_total=float(tx['Amount'].sum())

    r=start_row
    write_sec_hdr(ws,r,f'▸  Spending By Team  ({len(team)} Teams)',GREEN,6,9); r+=1
    write_col_hdrs(ws,r,['Team','Total Spent ($)','# Trans','% Of Total'],'2E8B3B',6); r+=1

    for i,row in team.iterrows():
        fill=ALT_TEAM if i%2==0 else WHITE; pct=row['Total Spent']/grand_total if grand_total else 0
        for j,(val,fmt,aln) in enumerate(zip([row['Team'],row['Total Spent'],int(row['Transactions']),pct],
                                              [None,'$#,##0.00','#,##0','0.0%'],['left','center','center','center'])):
            c=ws.cell(row=r,column=6+j,value=val); c.font=cfont(); c.fill=PatternFill('solid',fgColor=fill)
            c.border=bold_left_border() if j==0 else tborder()
            c.alignment=Alignment(horizontal=aln,vertical='center')
            if fmt: c.number_format=fmt
        ws.row_dimensions[r].height=18; r+=1

    for j,(val,fmt,aln) in enumerate(zip(['Grand Total',grand_total,int(team['Transactions'].sum()),1.0],
                                          [None,'$#,##0.00','#,##0','0.0%'],['left','center','center','center'])):
        c=ws.cell(row=r,column=6+j,value=val); c.font=Font(name='Arial',bold=True,color=WHITE,size=11)
        c.fill=PatternFill('solid',fgColor=DARK_BLUE); c.border=tborder()
        c.alignment=Alignment(horizontal=aln,vertical='center')
        if fmt: c.number_format=fmt
    ws.row_dimensions[r].height=24; r+=1
    return r

# ── COMPANY TAB ───────────────────────────────────────────────────────────────

def build_company_tab(wb,tx,report_title,actual_date_range,tab_name,is_first=False):
    ws=wb.active if is_first else wb.create_sheet(tab_name[:31])
    if is_first: ws.title=tab_name[:31]
    subtitle=f'{actual_date_range}  |  {len(tx):,} Transactions  |  ${tx["Amount"].sum():,.2f} Total'
    write_title(ws,report_title,subtitle,9)
    acct_end=write_account_summary(ws,tx,4)
    write_team_summary(ws,tx,4)
    widths={1:24,2:16,3:10,4:12,5:3,6:30,7:16,8:10,9:12}
    for col,w in widths.items(): ws.column_dimensions[get_column_letter(col)].width=w
    for row in range(4,acct_end+80):
        c=ws.cell(row=row,column=5,value=''); c.fill=PatternFill('solid',fgColor='F0F0F0')
    ws.freeze_panes='A4'

# ── TRANSACTIONS TAB ──────────────────────────────────────────────────────────

def build_transactions_tab(wb,tx,report_title,actual_date_range):
    tdf=tx.sort_values('Amount',ascending=False).reset_index(drop=True)
    ws=wb.create_sheet('Transactions')
    write_title(ws,report_title,f'{actual_date_range}  |  {len(tdf):,} Transactions',6)
    for col,h in enumerate(['Date','Type','Name','Description','Account','Amount ($)'],1):
        c=ws.cell(row=4,column=col,value=h); c.font=hfont(); c.fill=PatternFill('solid',fgColor=MED_BLUE)
        c.alignment=Alignment(horizontal='center',vertical='center'); c.border=tborder()
    ws.row_dimensions[4].height=22
    for i,row in tdf.iterrows():
        r=i+5; fill=ALT_ROW if i%2==0 else WHITE
        vals=[row['Date'].strftime('%m/%d/%Y') if pd.notna(row['Date']) else '',row['Type'],row['Name'],row['Description'],row['Account'],row['Amount']]
        for col,val in enumerate(vals,1):
            c=ws.cell(row=r,column=col,value=val); c.font=cfont(sz=9)
            c.fill=PatternFill('solid',fgColor=fill); c.border=tborder(); c.alignment=Alignment(horizontal='left',vertical='center')
        ws.cell(row=r,column=6).alignment=Alignment(horizontal='right',vertical='center')
        ws.cell(row=r,column=6).number_format='$#,##0.00'; ws.row_dimensions[r].height=16
    for i,w in enumerate([13,16,22,52,18,14],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes='A5'

# ── PIPELINE ──────────────────────────────────────────────────────────────────

def generate_report_bytes(file_bytes):
    df=pd.read_excel(io.BytesIO(file_bytes),header=None)
    valid,err=validate_gl(df)
    if not valid: raise ValueError(err)
    fmt=detect_format(df)
    company_title,actual_label,next_label,report_title,filename=build_meta(df,fmt)
    tx=load_transactions(df,fmt)
    wb=Workbook()

    if fmt=='consolidated':
        companies=sorted(tx['Company'].dropna().unique(),key=lambda x:tx[tx['Company']==x]['Amount'].sum(),reverse=True)
        all_title=f'All Companies - Installment Projections - {next_label}'
        build_company_tab(wb,tx,all_title,actual_label,'All Companies',is_first=True)
        for company in companies:
            co_tx=tx[tx['Company']==company].copy()
            co_title=f'{smart_title(company)} - Installment Projections - {next_label}'
            build_company_tab(wb,co_tx,co_title,actual_label,smart_title(company))
        build_transactions_tab(wb,tx,all_title,actual_label)
        stats={'transactions':len(tx),'accounts':tx['Account'].nunique(),'teams':tx['Name'].nunique(),
               'total':float(tx['Amount'].sum()),'date_range':actual_label,'companies':len(companies)}
    else:
        build_company_tab(wb,tx,report_title,actual_label,'Summary',is_first=True)
        build_transactions_tab(wb,tx,report_title,actual_label)
        stats={'transactions':len(tx),'accounts':tx['Account'].nunique(),'teams':tx['Name'].nunique(),
               'total':float(tx['Amount'].sum()),'date_range':actual_label,'companies':1}

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return filename,buf.read(),stats

# ── UI ────────────────────────────────────────────────────────────────────────

st.markdown("""
<div class="hero">
    <h1>🎟️ Installment<br><em>Projections</em></h1>
    <p>Drop in a General Ledger export and get your formatted report instantly.</p>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="upload-card">', unsafe_allow_html=True)
st.markdown('<div class="step-label">Step 1 — Upload your file(s)</div>', unsafe_allow_html=True)

uploaded_files=st.file_uploader(label="General Ledger (.xlsx)",type=["xlsx"],accept_multiple_files=True,label_visibility="collapsed")

if uploaded_files:
    st.markdown('<hr class="divider">', unsafe_allow_html=True)
    n=len(uploaded_files)
    st.markdown(f'<div class="step-label">Step 2 — Your {"report" if n==1 else f"{n} reports"}</div>', unsafe_allow_html=True)

    for uploaded in uploaded_files:
        with st.spinner(f"Building report for {uploaded.name}…"):
            try:
                filename,xlsx_bytes,stats=generate_report_bytes(uploaded.read())
                co_str=f'{stats["companies"]} companies · ' if stats["companies"]>1 else ''
                st.markdown(f"""
                <div class="success-box">
                    <p class="file-title">📄 {filename}</p>
                    <p class="file-meta">{stats['date_range']} · {co_str}{stats['transactions']:,} transactions</p>
                </div>""", unsafe_allow_html=True)
                st.markdown(f"""
                <div class="stats-row">
                    <div class="stat-card"><div class="stat-num">{stats['transactions']:,}</div><div class="stat-lbl">Transactions</div></div>
                    <div class="stat-card"><div class="stat-num">{stats['accounts']}</div><div class="stat-lbl">Accounts</div></div>
                    <div class="stat-card"><div class="stat-num">{stats['teams']}</div><div class="stat-lbl">Teams</div></div>
                    <div class="stat-card"><div class="stat-num">${stats['total']:,.0f}</div><div class="stat-lbl">Total Spent</div></div>
                </div>""", unsafe_allow_html=True)
                st.download_button(label=f"⬇️  Download — {filename}",data=xlsx_bytes,file_name=filename,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key=filename)
            except Exception as e:
                st.markdown(f'<div class="err-box">⚠️ <strong>{uploaded.name}</strong>: {e}</div>', unsafe_allow_html=True)
        st.markdown('<hr class="divider">', unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)
st.markdown('<div class="footer">Single & consolidated GL · Multiple files · One tab per company · No circular references</div>', unsafe_allow_html=True)
